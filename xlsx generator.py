"""
Génération de factures au format Excel (.xlsx) pour Swanky Apartments PMS.
Utilise openpyxl.
"""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BLEU_FONCE = "182236"
GRIS_CLAIR = "F0F0F5"


def _fmt(v):
    return f"{v:,.0f}".replace(",", " ") + " FCFA"


def generer_xlsx_facture(facture: dict, nom_etablissement="Swanky Apartments") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Facture"

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 28

    # --- En-tête ---
    ws.merge_cells("A1:B1")
    ws["A1"] = nom_etablissement
    ws["A1"].font = Font(bold=True, size=16, color=BLEU_FONCE)

    ws.merge_cells("A2:B2")
    ws["A2"] = "Reçu de facturation"
    ws["A2"].font = Font(size=11, italic=True, color="606060")

    # --- Infos ---
    infos = [
        ("Facture N°", facture["numero_facture"]),
        ("Date", facture["date_creation"][:16]),
        ("Client", facture["client"]),
        ("Chambre", f"{facture['chambre_nom']} ({facture.get('chambre_type', '')})"),
        ("Mode de paiement", facture.get("mode_paiement", "Espèces")),
        ("Statut", facture.get("statut_paiement", "Payé")),
    ]
    row = 4
    for label, valeur in infos:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=valeur)
        row += 1

    row += 1
    # --- Tableau détail ---
    entetes = ["Nuitée(s)", "Tarif unitaire (TTC)", "Total HT", f"TVA ({facture['taux_tva']:g}%)", "Total TTC"]
    for i, texte in enumerate(entetes):
        cell = ws.cell(row=row, column=i + 1, value=texte)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=BLEU_FONCE)
        cell.alignment = Alignment(horizontal="center")

    row += 1
    valeurs = [
        facture["nuitees"],
        facture["tarif_unitaire"],
        facture["montant_ht"],
        facture["montant_tva"],
        facture["montant_ttc"],
    ]
    for i, v in enumerate(valeurs):
        cell = ws.cell(row=row, column=i + 1, value=v)
        cell.number_format = "#,##0"
        cell.alignment = Alignment(horizontal="center")

    row += 2
    ws.cell(row=row, column=4, value="TOTAL TTC :").font = Font(bold=True, size=12)
    total_cell = ws.cell(row=row, column=5, value=facture["montant_ttc"])
    total_cell.font = Font(bold=True, size=12, color="FFFFFF")
    total_cell.fill = PatternFill("solid", fgColor=BLEU_FONCE)
    total_cell.number_format = "#,##0"

    if facture.get("notes"):
        row += 2
        ws.cell(row=row, column=1, value="Notes :").font = Font(italic=True)
        ws.cell(row=row, column=2, value=facture["notes"]).font = Font(italic=True)

    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 20

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def generer_xlsx_historique(factures: list, nom_etablissement="Swanky Apartments") -> bytes:
    """Exporte la liste complète des factures dans un seul classeur Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Historique factures"

    entetes = ["N° Facture", "Date", "Chambre", "Type", "Client", "Nuitées",
               "Tarif unitaire", "Total HT", "TVA", "Total TTC", "Mode paiement", "Statut"]
    for i, texte in enumerate(entetes):
        cell = ws.cell(row=1, column=i + 1, value=texte)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=BLEU_FONCE)

    for r, f in enumerate(factures, start=2):
        valeurs = [
            f["numero_facture"], f["date_creation"][:16], f["chambre_nom"], f.get("chambre_type", ""),
            f["client"], f["nuitees"], f["tarif_unitaire"], f["montant_ht"], f["montant_tva"],
            f["montant_ttc"], f.get("mode_paiement", ""), f.get("statut_paiement", ""),
        ]
        for c, v in enumerate(valeurs, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            if c in (7, 8, 9, 10):
                cell.number_format = "#,##0"

    for col in range(1, len(entetes) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()

